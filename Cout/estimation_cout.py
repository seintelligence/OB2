import os
from typing import Tuple, Optional
from openpyxl import load_workbook
import locale
import time

class ExcelInteraction:
    def __init__(self, file_path: str = "Cout/Modele Devis v1.xlsx"):
        self.file_path = os.path.abspath(file_path)
        self.workbook = None
        locale.setlocale(locale.LC_NUMERIC, 'C')

    def connect_to_excel(self) -> bool:
        try:
            print(f"Connection à Excel pour le fichier: {self.file_path}")
            self.workbook = load_workbook(self.file_path, data_only=True)
            
            print("\nNoms définis disponibles avec leurs références:")
            for name in self.workbook.defined_names:
                named_range = self.workbook.defined_names[name]
                destinations = list(named_range.destinations)
                print(f"- {name}: {destinations[0] if destinations else 'No destination'}")
            
            return True
        except Exception as e:
            print(f"Erreur lors de la connexion: {str(e)}")
            self.cleanup()
            return False

    def set_inputs_get_outputs(self, hauteur_mur: float, largeur_mur: float) -> Tuple[Optional[float], Optional[float]]:
        try:
            # Mise à jour des inputs
            ws_input = self.workbook['Input']
            print(f"\nÉcriture hauteur {hauteur_mur} en B2")
            ws_input['B2'] = float(str(hauteur_mur).replace(',', '.'))
            
            print(f"Écriture largeur {largeur_mur} en B3")
            ws_input['B3'] = float(str(largeur_mur).replace(',', '.'))
            
            # Calculer et afficher la surface pour information
            surface = (hauteur_mur * largeur_mur) / 10000
            print(f"Surface calculée: {surface} m²")
            
            # Sauvegarder et attendre
            print("Sauvegarde et temporisation...")
            self.workbook.save(self.file_path)
            time.sleep(1)  # Attendre 1 seconde
            
            # Recharger le fichier pour avoir les valeurs calculées
            self.workbook = load_workbook(self.file_path, data_only=True)
            ws_output = self.workbook['Output']
            
            # Lire les outputs
            cout_m2 = ws_output['B2'].value
            cout_mur = ws_output['B3'].value
            
            print(f"Valeurs calculées - Au m2: {cout_m2}, Total: {cout_mur}")
            
            if cout_m2 is None or cout_mur is None:
                raise ValueError("Valeurs non trouvées")
                
            return round(float(cout_m2), 2), round(float(cout_mur), 2)
            
        except Exception as e:
            print(f"Erreur lors du traitement: {str(e)}")
            return None, None

    def cleanup(self):
        try:
            if self.workbook:
                self.workbook.close()
        except Exception as e:
            print(f"Erreur lors du nettoyage: {str(e)}")

def process_wall_costs(hauteur_mur_cm: float, largeur_mur_cm: float) -> Tuple[Optional[float], Optional[float]]:
    """
    Traite les coûts du mur avec des dimensions en centimètres.
    """
    print(f"\nTraitement pour un mur de {hauteur_mur_cm}cm x {largeur_mur_cm}cm")
    excel_handler = None

    try:
        excel_handler = ExcelInteraction()
        
        if not excel_handler.connect_to_excel():
            raise Exception("Impossible de se connecter au fichier Excel")
            
        # Une seule fonction pour mettre à jour et récupérer les résultats
        cout_m2, cout_mur = excel_handler.set_inputs_get_outputs(hauteur_mur_cm, largeur_mur_cm)
        
        if cout_m2 is None or cout_mur is None:
            raise ValueError("Impossible de récupérer les coûts calculés")
            
        print(f"Traitement terminé avec succès : {cout_m2} EUR/m2 - {cout_mur} EUR total")
        return float(cout_m2), float(cout_mur)
            
    except Exception as e:
        print(f"Erreur lors du traitement: {str(e)}")
        return None, None
        
    finally:
        if excel_handler:
            excel_handler.cleanup()